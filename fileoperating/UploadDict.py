from openpyxl import load_workbook
import re
import json


def get_info_data(file_path):
    """
    获取info标注文本信息
    :param file_path:info.xlsx文件路径
    :return:data_list
    """
    wb = load_workbook(file_path)
    sheet = wb.worksheets[0]
    contactList = []
    music_list=[]
    count = 2
    row = sheet.max_row + 1
    for item in sheet["A"][1:row]:
        if sheet["G" + str(count)].value == "关键字":
            if sheet["E" + str(count)].value != None:
                pthon_name = sheet["E" + str(count)].value
            else:
                pthon_name = sheet["D" + str(count)].value
            if pthon_name != None:
                pthonList = {"name": str(pthon_name)}
                contactList.append(pthonList)
        elif sheet["G" + str(count)].value == "关键字":
            if sheet["E" + str(count)].value != None:
                music_name = sheet["E" + str(count)].value
            else:
                music_name = sheet["D" + str(count)].value
            if music_name != None:
                musiclist = {"name": str(music_name)}
                music_list.append(musiclist)
        count += 1
    list = dict(contactList=contactList, music_list=music_list)
    json.dumps(list, ensure_ascii=False)
    print(len(list['contactList']))
    print(len(list['music_list']))
    print(json.dumps(list, ensure_ascii=False))
    return list


def Export_txt(txt_path,data_list):
      with open(txt_path, "w+", encoding="utf-8", errors="ignore") as f:
            json.dump(data_list,f)


if __name__ =="__main__":
    file_path=r"automationdemo.github.io\fileoperating\test.xlsx"
    txt_path=r"automationdemo.github.io\fileoperating\test_dicts.txt"
    list=get_info_data(file_path)
    Export_txt(txt_path,list)